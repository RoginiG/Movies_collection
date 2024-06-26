import mysql.connector
import openpyxl
from openpyxl import workbook,load_workbook
data=mysql.connector.connect(host="localhost",username="root",password="12345",database="movie_database")
cur=data.cursor()
def movie_data():
    loc=("C:\\Users\\rogin\\OneDrive\\Documents\\movie.xlsx")
    l=list()
    a=openpyxl.load_workbook(loc)
    sheet=a.worksheets[0]
    for i in range(1, sheet.max_row + 1):
        col1_value = sheet.cell(row=i, column=1).value
        col2_value = sheet.cell(row=i, column=2).value
        col3_value = sheet.cell(row=i, column=3).value
        col4_value = sheet.cell(row=i, column=4).value
        value_tuples=(col1_value, col2_value, col3_value, col4_value)
        l.append(value_tuples)
    sql="insert into movie_collection values(%s,%s,%s,%s)"
    cur.executemany(sql,l)
    data.commit()
    data.close()
movie_data()
def movie_data_view():
    cur.execute("select*from movie_collection")
    myresult=cur.fetchall()
    for i in myresult:
        print (i)
    myresult=cur.fetchall()
    print("Access the particular row")
    cur.execute("select*from movie_collection where movie='mouna ragam'")
    myresult=cur.fetchall()
    for row in myresult:
        print(row)
movie_data_view()
